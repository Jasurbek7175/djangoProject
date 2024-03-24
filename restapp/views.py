from django.shortcuts import render
# Create your views here.
from django.shortcuts import render, redirect
from .forms import UploadFileForm
from .models import YoursDataModel
import pandas as pd
from .serializers import ModelSerializer
from django.core.serializers import serialize
from datetime import datetime
from django.utils import timezone
from rest_framework.views import APIView
from rest_framework import response

from django.shortcuts import render, HttpResponse
from translate import Translator


# Create your views here.

def translation(request):
    if request.method == "POST":
        text = request.POST["translate"]
        language = request.POST["language"]
        translator = Translator(to_lang=language)
        translation = translator.translate(text)
        return HttpResponse(translation)
    return render(request, "index.html")

class LoanInfoView(APIView):
    def get(self, request, *args, **kwargs):
        serializer = YourDataModelSerializer(data=request.query_params)
        if serializer.is_valid():
            name = serializer.validated_data['name']
            age = serializer.validated_data['age']
            print('1515151515151', age)
            person = YoursDataModel.objects.get(name=name).values()
            # Check if name is "Maruf" and age is 119
            if person:
                # Assuming there is a person in your database with name "Maruf" and age 119
                serializer = YourDataModelSerializer(person, many=True)
                return Response(serializer.data)
            else:
                return Response({'error': 'Invalid parameters'}, status=400)
        else:
            return Response(serializer.errors, status=400)


from .tasks import slow_func


def index(request):
    slow_func.delay(13244515)
    return HttpResponse('Site is working!')


def get_objects_in_month(year):
    start_date = timezone.make_aware(datetime(2023, 10, 1))
    end_date = timezone.make_aware(datetime(2023, 10, 31))  # Next month's 1st day

    objects_in_month = YoursDataModel.objects.filter(
        created_time__gte=start_date,
        created_time__lt=end_date
    )

    return objects_in_month


from openpyxl import load_workbook
from django.conf import settings
from django.core.files.storage import FileSystemStorage


def import_from_excel(request):
    if request.method == 'POST':
        excel_file = request.FILES['excel_file']
        wb = load_workbook(excel_file)
        ws = wb.active

        folder = 'media/'
        fs = FileSystemStorage(location=folder)  # defaults to   MEDIA_ROOT
        filename = fs.save(excel_file.name, excel_file)
        file_url = fs.url(filename)

        # Create a list to store all instances
        instances = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                contract_date = row[3].date()
                a = YoursDataModel.objects.create(
                    name=row[0],  # Assuming Name is the first column in your Excel file
                    age=row[1],  # Assuming Age is the second column
                    city=row[2],
                    contract_date=contract_date
                    # Assuming City is the third column
                    # Add more fields as needed
                )

                # Append the created instance to the list
                instances.append(a)
            except Exception as e:
                print(f"Error processing row {row}: {e}")

        # Serialize all instances
        serializer = YourDataModelSerializer(instances, many=True)
        serialized_data = serializer.data

        # Convert the serialized data to JSON
        data = json.dumps(serialized_data)

        # Define the URL for the external API
        external_api_url = "http://127.0.0.1:8000/endpoint/"

        # Set headers for the external API request
        external_api_headers = {
            'Content-Type': 'application/json'
        }

        # Make a POST request to the external API
        external_api_response = requests.post(
            external_api_url,
            headers=external_api_headers,
            data=data
        )

        # Extract relevant information from the API response
        response_data = {
            'status_code': external_api_response.status_code,
            'content': external_api_response.json(),
        }

        # if external_api_response.status_code == 200:
        #     return redirect('success')  # Replace with your actual URL name or path

        # Return the response as a JSON object
        # return render(request, 'success.html', response_data)

        return redirect('success')

    return render(request, 'upload_file.html')


# def import_from_excel(request):
#     if request.method == 'POST':
#         excel_file = request.FILES['excel_file']
#         wb = load_workbook(excel_file)
#         ws = wb.active
#
#         for row in ws.iter_rows(min_row=2, values_only=True):
#             a = YoursDataModel.objects.create(
#                 name=row[0],  # Assuming Name is the first column in your Excel file
#                 age=row[1],  # Assuming Age is the second column
#                 city=row[2],  # Assuming City is the third column
#                 # Add more fields as needed
#             )
#
#         # Serialize the created instances
#         serializer = YourDataModelSerializer(YoursDataModel.objects.all(), many=True)
#         serialized_data = serializer.data
#
#         # Convert the serialized data to JSON
#         data = json.dumps(serialized_data)
#
#         # Define the URL for the external API
#         external_api_url = "http://127.0.0.1:8000/endpoint/"
#
#         # Set headers for the external API request
#         external_api_headers = {
#             'Content-Type': 'application/json'
#         }
#
#         # Make a POST request to the external API
#         external_api_response = requests.post(
#             external_api_url,
#             headers=external_api_headers,
#             data=data
#         )
#
#         # Extract relevant information from the API response
#         response_data = {
#             'status_code': external_api_response.status_code,
#             'content': external_api_response.text,
#         }
#
#         # Return the response as a JSON object
#         return JsonResponse(response_data, safe=False)
#
#     return render(request, 'upload_file.html')


# def import_from_excel(request):
#     if request.method == 'POST':
#         excel_file = request.FILES['excel_file']
#         wb = load_workbook(excel_file)
#         ws = wb.active
#
#         for row in ws.iter_rows(min_row=2, values_only=True):
#             a = YoursDataModel.objects.create(
#                 name=row['Name'],
#                 age=row['Age'],
#                 city=row['City'],
#                 # Add more fields as needed
#             )
#
#
#
#         v = YourDataModelSerializer(a, many=True).data
#
#         data = json.dumps(v)
#
#         # Define the URL for the external API
#         external_api_url = "http://127.0.0.1:8000/endpoint/"
#
#         # Set headers for the external API request
#         external_api_headers = {
#             # 'Authorization': 'Token aXBvdGVrYS1wZXJjZW50YWdlLXBheW1lbnQ6OzNpSGI3aTswNjRd',
#             'Content-Type': 'application/json'
#         }
#
#         # Make a POST request to the external API
#         external_api_response = requests.post(
#             external_api_url,
#             headers=external_api_headers,
#             # verify=False,
#             data=data
#         )
#
#         # print(data, external_api_response)
#
#         # Return the response from the external API
#         return JsonResponse(external_api_response)
#     return render(request, 'upload_file.html')


def handle_uploaded_file(file):
    df = pd.read_excel(file)
    for index, row in df.iterrows():
        YoursDataModel.objects.create(
            name=row['Name'],
            age=row['Age'],
            city=row['City'],
            # Add more fields as needed
        )
        # w = YoursDataModel.objects.all()
        #
        # v = YourDataModelSerializer(w, many=True).data
        #
        # data = json.dumps(v)
        #
        # # Define the URL for the external API
        # external_api_url = "https://192.168.254.217/front/api/houseconstruction/task/create"
        #
        # # Set headers for the external API request
        # external_api_headers = {
        #     'Authorization': 'Token aXBvdGVrYS1wZXJjZW50YWdlLXBheW1lbnQ6OzNpSGI3aTswNjRd',
        #     'Content-Type': 'application/json'
        # }
        #
        # # Make a POST request to the external API
        # external_api_response = requests.post(
        #     external_api_url,
        #     headers=external_api_headers,
        #     verify=False,  # Disabling SSL verification for simplicity; don't use in production
        #     data=data
        # )
        #
        # # Return the response from the external API
        # return JsonResponse(external_api_response.json())


def upload_file(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            handle_uploaded_file(request.FILES['file'])
            return redirect('success')
    else:
        form = UploadFileForm()
    return render(request, 'upload_file.html', {'form': form})


from rest_framework.renderers import JSONRenderer


def success(request):
    # Add any success message or redirection logic here
    # w = YoursDataModel.objects.all()

    # v = YourDataModelSerializer(w, many=True).data

    # data = json.dumps(v)

    # Define the URL for the external API
    # external_api_url = "http://127.0.0.1:8000/endpoint/"

    # Set headers for the external API request
    # external_api_headers = {
    #     # 'Authorization': 'Token aXBvdGVrYS1wZXJjZW50YWdlLXBheW1lbnQ6OzNpSGI3aTswNjRd',
    #     'Content-Type': 'application/json'
    # }

    # Make a POST request to the external API
    # external_api_response = requests.post(
    #     external_api_url,
    #     headers=external_api_headers,
    #     # verify=False,
    #     data=data
    # )

    # print(data, external_api_response)

    # Return the response from the external API
    # return JsonResponse(external_api_response)

    return render(request, 'success.html')


from django.http import HttpResponse
import pandas as pd
import json
from django.http import HttpResponse, JsonResponse


def create_excel_file(request):
    # Your data retrieval logic goes here
    data = {
        'Name': ['John', 'Jane', 'Bob'],
        'Age': [25, 30, 22],
        'City': ['New York', 'London', 'Paris'],
    }

    # Create a DataFrame from the data
    df = pd.DataFrame(data)

    # Create a response object with the appropriate content type for Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=my_project_data.xlsx'

    # Write the DataFrame to the response object
    df.to_excel(response, index=False, engine='openpyxl')

    return response


import requests
from rest_framework.views import APIView


# class Data():
def post_data(request):
    data1 = YoursDataModel.objects.all()

    df = pd.DataFrame(data1)

    # Convert the DataFrame to JSON format
    json_data = df.to_json(orient='records')

    # Parse the JSON data to a Python list
    parsed_data = json.loads(json_data)

    data = ModelSerializer(data=json_data, many=True)

    # Convert the updated array back to JSON
    updated_json_data = json.dumps(data)

    api_url = 'http://127.0.0.1:8000/endpoint/'
    headers = {'Content-Type': 'application/json'}
    response = requests.post(api_url, data=updated_json_data, headers=headers)

    # print(response)

    if response.status_code == 200:
        return JsonResponse({'status': 'Data successfully sent to the API'})
    else:
        return JsonResponse({'status': 'Failed to send data to the API'}, status=500)

    # For simplicity, returning the JSON as a response in this example
    # return JsonResponse({'updated_data': updated_json_data, 'data': json_data})


from rest_framework.views import APIView


class YourDataModelAPIView(APIView):
    def post(self, request, *args, **kwargs):
        data = request.data
        serializer = YourDataModelSerializer(data=data, many=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class DataModelAPIView(APIView):
    def post(self, request, *args, **kwargs):
        try:
            # Parse the incoming JSON data
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return Response({'error': 'Invalid JSON data'}, status=status.HTTP_400_BAD_REQUEST)

        # Create and save objects for each item in the array
        for item in data:
            serializer = YourDataModelSerializer(data=item)
            if serializer.is_valid():
                serializer.save()
            else:
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        return Response({'status': 'Data successfully saved'}, status=status.HTTP_201_CREATED)


from django.db.models.signals import post_save
from django.dispatch import receiver
import requests

from .models import YoursDataModel


@receiver(post_save, sender=YoursDataModel)
def send_data_to_api(sender, instance, **kwargs):
    api_url = 'http://127.0.0.1:8000/endpoint/'  # Replace with your actual API endpoint
    headers = {'Content-Type': 'application/json'}

    # data_to_send = {
    #     "Name": instance.name,
    #     "Age": instance.age,
    #     "City": instance.city,
    # }

    data = YoursDataModel.objects.all()

    df = pd.DataFrame(data)

    columns_to_exclude = ['id']
    df = df.loc[:, (df != 0).any(axis=0) & ~df.columns.isin(columns_to_exclude)]

    # Convert the DataFrame to JSON format
    json_data = df.to_json(orient='records')

    # Parse the JSON data to a Python list
    parsed_data = json.loads(json_data)

    # Append the new object to the array
    # new_object = {'Name': 'Alice', 'Age': 28, 'City': 'Berlin'}
    # parsed_data.append(new_object)

    # Convert the updated array back to JSON
    updated_json_data = json.dumps(parsed_data)

    try:
        response = requests.post(api_url, json=updated_json_data, headers=headers)

        if response.status_code == 201:
            print('Data successfully sent to the API')
        else:
            print(f'Failed to send data to the API. Response code: {response.status_code}')

    except requests.RequestException as e:
        print(f'Error during request: {str(e)}')


from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView
import json
from .serializers import YourDataModelSerializer


class ModelAPIView(APIView):
    def post(self, request, *args, **kwargs):
        try:
            # Parse the incoming JSON data
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return Response({'error': 'Invalid JSON data'}, status=status.HTTP_400_BAD_REQUEST)

        # Create and save objects for each item in the array
        for item in data:
            serializer = YourDataModelSerializer(data=item)
            if serializer.is_valid():
                serializer.save()
            else:
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        return Response({'status': 'Data successfully saved'}, status=status.HTTP_201_CREATED)


from rest_framework import status
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.response import Response
from rest_framework.views import APIView

from .serializers import FileUploadSerializer


class FileUploadAPIView(APIView):
    parser_classes = (MultiPartParser, FormParser)
    serializer_class = FileUploadSerializer

    def get(self, request):
        a = UploadedFile.objects.all()
        b = FileUploadSerializer(a, many=True)
        return Response(b.data)

    def post(self, request, *args, **kwargs):
        serializer = self.serializer_class(data=request.data, many=True)
        print(serializer)
        if serializer.is_valid():
            # you can access the file like this from serializer
            # uploaded_file = serializer.validated_data["file"]
            serializer.save()
            return Response(
                serializer.data,
                status=status.HTTP_201_CREATED
            )

        return Response(
            serializer.errors,
            status=status.HTTP_400_BAD_REQUEST
        )


from .models import UploadedFile


def uploadpdf(request):
    file = UploadedFile.objects.all()
    templates = 'success.html'
    context = {
        'file': file
    }
    return render(request, template_name=templates, context=context)


import os
from django.shortcuts import get_object_or_404
from wsgiref.util import FileWrapper


def download_file(request, file_id):
    uploaded_file = get_object_or_404(UploadedFile, pk=file_id)

    # Get the file path
    file_path = os.path.join(settings.MEDIA_ROOT, str(uploaded_file.file))

    # Open the file for reading
    with open(file_path, 'rb') as file:
        response = HttpResponse(FileWrapper(file), content_type='application/octet-stream')

        # Set the content-disposition header for the browser to download the file
        response['Content-Disposition'] = f'attachment; filename="{os.path.basename(file_path)}"'

        return response
